import os
import json
import csv
import openpyxl

class DataAggregator:
    def __init__(self, folder_path, valid_subfolders=None):
        self.folder_path = folder_path
        self.valid_subfolders = valid_subfolders
        self.cached_json_map = None

    def get_indexed_jsons(self):
        if self.cached_json_map is not None:
            return self.cached_json_map
        
        json_map = {}

        for subfolder in os.listdir(self.folder_path):
            subfolder_path = os.path.join(self.folder_path, subfolder, "taskgraph")

            if self.valid_subfolders is None or (os.path.isdir(subfolder_path) and subfolder in self.valid_subfolders):
                json_name = subfolder + '_task_graph_metrics.json'
                json_file_path = os.path.join(subfolder_path, json_name)

                if os.path.exists(json_file_path):
                    with open(json_file_path, 'r') as json_file:
                        data = json.load(json_file)

                        if 'appName' in data:
                            app_name = data['appName']
                            json_map[app_name] = data

        self.cached_json_map = json_map
        return json_map

    def output_combined_json(self, output_file_path='combined_output.json'):
        json_map = self.get_indexed_jsons()

        with open(output_file_path, 'w') as output_file:
            json.dump(json_map, output_file, indent=4)


    def output_excel_from_csv_list(self, csv_files, excel_filename='combined_output.xlsx', delete_csv=False):
        workbook = openpyxl.Workbook()

        for csv_file in csv_files:
            sheet_name = csv_file.split('.')[0]
            sheet = workbook.create_sheet(title=sheet_name)

            with open(csv_file, 'r') as file:
                csv_reader = csv.reader(file)
                for row_index, row in enumerate(csv_reader, start=1):
                    for col_index, value in enumerate(row, start=1):
                        sheet.cell(row=row_index, column=col_index, value=value)
            if delete_csv:
                os.remove(csv_file)

        workbook.remove(workbook['Sheet'])
        workbook.save(excel_filename)


    def output_general_stats(self, csv_file_path='general_stats.csv'):
        json_map = self.get_indexed_jsons()

        with open(csv_file_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)

            header = [
                "Suite",
                "Benchmark", 
                "regularTasks",
                "externalTasks",
                "inlinableCalls",
                "globalVariables",
            ]
            csv_writer.writerow(header)

            # Write rows
            for app_name, data in json_map.items():
                row_data = [
                    app_name.split('-')[0],
                    '-'.join(app_name.split('-')[1:-1]),

                    data.get('counts', {}).get('regularTasks', 'N/A'),
                    data.get('counts', {}).get('externalTasks', 'N/A'),
                    data.get('counts', {}).get('inlinableCalls', 'N/A'),
                    data.get('counts', {}).get('globalVars', 'N/A'),
                ]
                csv_writer.writerow(row_data)
        return csv_file_path    
        

    def output_unique_task_data(self, csv_file_path='unique_task_data.csv'):
        json_map = self.get_indexed_jsons()

        with open(csv_file_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)

            header = [
                "Suite",
                "Benchmark", 
                "regularTasks",
                "externalTasks",
                "inlinableCalls",
                "globalVariables",
            ]
            csv_writer.writerow(header)

            # Write rows
            for app_name, data in json_map.items():
                row_data = [
                    app_name.split('-')[0],
                    '-'.join(app_name.split('-')[1:-1]),

                    data.get('counts', {}).get('regularTasks', 'N/A'),
                    data.get('counts', {}).get('externalTasks', 'N/A'),
                    data.get('counts', {}).get('inlinableCalls', 'N/A'),
                    data.get('counts', {}).get('globalVars', 'N/A'),
                ]
                csv_writer.writerow(row_data)
        return csv_file_path

    
    def output_data_per_task(self, csv_file_path='data_per_task.csv'):
        json_map = self.get_indexed_jsons()

        with open(csv_file_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)

            header = [
                "Suite",
                "Benchmark", 
            ]
            csv_writer.writerow(header)

            # Write rows
            for app_name, data in json_map.items():
                row_data = [
                    app_name.split('-')[0],
                    '-'.join(app_name.split('-')[1:-1]),

                ]
                csv_writer.writerow(row_data)
        return csv_file_path
    

    def output_global_var_data(self, csv_file_path='global_var_data.csv'):
        json_map = self.get_indexed_jsons()

        with open(csv_file_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)

            header = [
                "Suite",
                "Benchmark", 
            ]
            csv_writer.writerow(header)

            # Write rows
            for app_name, data in json_map.items():
                row_data = [
                    app_name.split('-')[0],
                    '-'.join(app_name.split('-')[1:-1]),

                ]
                csv_writer.writerow(row_data)
        return csv_file_path
    

    def output_data_source_distance(self, csv_file_path='data_source_distance.csv'):
        json_map = self.get_indexed_jsons()

        with open(csv_file_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)

            header = [
                "Suite",
                "Benchmark", 
            ]
            csv_writer.writerow(header)

            # Write rows
            for app_name, data in json_map.items():
                row_data = [
                    app_name.split('-')[0],
                    '-'.join(app_name.split('-')[1:-1]),

                ]
                csv_writer.writerow(row_data)
        return csv_file_path
    

    def output_parallel_tasks(self, csv_file_path='parallel_tasks.csv'):
        json_map = self.get_indexed_jsons()

        with open(csv_file_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)

            header = [
                "Suite",
                "Benchmark", 
            ]
            csv_writer.writerow(header)

            # Write rows
            for app_name, data in json_map.items():
                row_data = [
                    app_name.split('-')[0],
                    '-'.join(app_name.split('-')[1:-1]),

                ]
                csv_writer.writerow(row_data)
        return csv_file_path