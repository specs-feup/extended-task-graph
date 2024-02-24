import os
import csv
import time
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, PatternFill, Border, Side


class ExcelConverter:

    def csv_files_to_excel(
        self,
        csv_files,
        excel_filename="combined_output",
        delete_csv=False,
        ranges_for_merging={},
    ):
        workbook = openpyxl.Workbook()

        for csv_file in csv_files:
            if os.path.getsize(csv_file) > 3:
                sheet_name = csv_file.split(".")[0].split("/")[-1].split("\\")[-1]
                sheet = workbook.create_sheet(title=sheet_name)

                with open(csv_file, "r") as file:
                    csv_reader = csv.reader(file)
                    for row_index, row in enumerate(csv_reader, start=1):
                        for col_index, value in enumerate(row, start=1):
                            sheet.cell(row=row_index, column=col_index, value=value)

            if delete_csv:
                os.remove(csv_file)
                pass

            merge_range = ranges_for_merging.get(sheet_name, range(0, 0))
            # self.create_styles(workbook)
            # self.convert_numbers(sheet)
            self.merge_and_center(sheet, merge_range)
            self.center_and_auto_size(sheet)
            self.set_header_format(sheet)

        workbook.remove(workbook["Sheet"])
        actual_filename = self.generate_filename(excel_filename)
        workbook.save(actual_filename)

    def generate_filename(self, filename):
        if filename.endswith(".xlsx"):
            filename = filename[:-5]

        timestamp = int(time.time())
        filename = f"{filename}_{timestamp}.xlsx"
        return filename

    def create_styles(self, workbook):
        integer_style = NamedStyle(name="Integer")
        integer_style.number_format = "0"

        float_style = NamedStyle(name="Float")
        float_style.number_format = "0.00"

        percentage_style = NamedStyle(name="Percentage")
        percentage_style.number_format = "0.00%"

        try:
            workbook.add_named_style(integer_style)
            workbook.add_named_style(float_style)
            workbook.add_named_style(percentage_style)
        except ValueError:
            pass

    def convert_numbers(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                # Check column header for percentage symbol
                if cell.row == 1 and "%" in str(cell.value):
                    column_index = cell.column
                    for row_num in range(2, sheet.max_row + 1):
                        cell_in_column = sheet.cell(row=row_num, column=column_index)
                        try:
                            cell_in_column.value = (
                                float(cell_in_column.value.strip("%")) / 100
                            )
                            cell_in_column.style = "Percentage"
                        except (ValueError, AttributeError):
                            pass  # Value cannot be converted to float or doesn't contain '%', leave it unchanged
                else:
                    # Try converting to integer
                    try:
                        cell.value = int(cell.value)
                        cell.style = "Integer"
                    except (TypeError, ValueError):
                        # If conversion fails, try converting to float
                        try:
                            cell.value = round(float(cell.value), 2)
                            cell.style = "Float"
                        except (TypeError, ValueError):
                            pass  # Value cannot be converted to int or float, leave it unchanged

    def set_header_format(self, sheet):
        sheet.freeze_panes = "A2"

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
            )
            cell.border = Border(top=Side(style="thin"))

    def merge_and_center(self, sheet, column_range=range(1, 3)):
        max_row = sheet.max_row

        for col in column_range:  # Only consider the first two columns
            current_value = sheet.cell(row=1, column=col).value
            start_row = 1

            for row in range(2, max_row + 1):
                cell_value = sheet.cell(row=row, column=col).value

                if cell_value != current_value:
                    if start_row < row - 1:
                        # Merge cells if there is more than one consecutive cell with the same value
                        sheet.merge_cells(
                            start_row=start_row,
                            end_row=row - 1,
                            start_column=col,
                            end_column=col,
                        )

                        # Center text both horizontally and vertically within the merged cells
                        for merge_row in range(start_row, row):
                            sheet.cell(row=merge_row, column=col).alignment = (
                                openpyxl.styles.Alignment(
                                    horizontal="center",
                                    vertical="center",
                                    wrap_text=True,  # Wrap text for merged cells
                                )
                            )

                    current_value = cell_value
                    start_row = row

            # Merge the last group of cells in the column
            if start_row < max_row:
                sheet.merge_cells(
                    start_row=start_row,
                    end_row=max_row,
                    start_column=col,
                    end_column=col,
                )

                # Center text both horizontally and vertically within the merged cells
                for merge_row in range(start_row, max_row + 1):
                    sheet.cell(row=merge_row, column=col).alignment = (
                        openpyxl.styles.Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=True,  # Wrap text for merged cells
                        )
                    )

    def center_and_auto_size(self, sheet):
        # Center text and auto-size columns and rows to fit the content
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                try:
                    cell.alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        horizontal="center",
                        vertical="center",
                    )
                except:
                    pass

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
